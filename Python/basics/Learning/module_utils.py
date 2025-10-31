import random

#finding the max number
def find_max(number_list):
    max_number = 0
    for number in number_list:
        if number > max_number:
            max_number = number
    return max_number

def roll_dice(num):
    return random.randint(2,num)


#Processing a workbook
import openpyxl as xl
try:
    def process_workbook(filename):

        workbook = xl.load_workbook(filename)
        sheet = workbook['Sheet1']

        for row in range(2,sheet.max_row + 1):
            sheet.cell(1,4).value = 'Corrected_Revenue'
            revenue = sheet.cell(row,3)
            #print(revenue.value)
            correct_price = round(revenue.value * 0.9,2)
            #print(correct_price)
            corrected_price_cell = sheet.cell(row,4)
            #print(corrected_price_cell)
            corrected_price_cell.value = correct_price
            #print(corrected_price_cell.value)

        workbook.save(filename)
        print(f'{filename} processed successfully')
except FileNotFoundError:
    print('File does not exist')



#main file code
import Learning.module_utils as m

print(m.roll_dice(4))